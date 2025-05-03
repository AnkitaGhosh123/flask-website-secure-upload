import random
import os
from openpyxl import load_workbook, Workbook
from cryptography.fernet import Fernet, InvalidToken

# Simulated two-factor authentication function
def two_factor_auth():
    code = random.randint(100000, 999999)
    print(f"Your 2FA code is: {code}")  # In real use, this would be sent securely
    user_input = input("Enter the 2FA code: ")
    return user_input == str(code)

# Device check function
def device_check(allowed_devices):
    current_device = "device1"  # Change this if needed
    return current_device in allowed_devices

# Placeholder Merkle root creation function
def create_merkle_root(data_list):
    return "MerkleRoot"  # Simulated placeholder

# Ledger update function (simple print)
def update_ledger(action):
    print(f"Ledger updated with action: {action}")

# Read from Excel file
def read_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return []

# Find first empty row in Excel sheet
def find_first_empty_row(sheet):
    for row in range(1, sheet.max_row + 2):
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1

# Write to Excel file
def write_to_excel_file(file_path, data):
    try:
        # Load workbook or create if it doesn't exist
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        sheet = workbook.active
        first_empty_row = find_first_empty_row(sheet)

        # Append new data
        for index, row in enumerate(data):
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=first_empty_row + index, column=col_num, value=value)

        workbook.save(file_path)
        print(f"Data written to {file_path}: {data}")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")

# Encrypt file
def encrypt_file(file_name, key):
    fernet = Fernet(key)
    with open(file_name, 'rb') as file:
        original = file.read()
    encrypted = fernet.encrypt(original)
    with open(file_name, 'wb') as encrypted_file:
        encrypted_file.write(encrypted)

# Decrypt file
def decrypt_file(file_name, key):
    fernet = Fernet(key)
    try:
        with open(file_name, 'rb') as encrypted_file:
            encrypted = encrypted_file.read()
        decrypted = fernet.decrypt(encrypted)
        with open(file_name, 'wb') as decrypted_file:
            decrypted_file.write(decrypted)
    except InvalidToken:
        print("Invalid decryption key! File could not be decrypted.")

# Generate encryption key
def generate_key():
    return Fernet.generate_key()

# Save key to file
def save_key(key, key_file):
    with open(key_file, 'wb') as kf:
        kf.write(key)

# Load key from file
def load_key(key_file):
    with open(key_file, 'rb') as kf:
        return kf.read()

# Main secure file function
def secure_file(file_path, allowed_devices, data_list, key_file):
    if not two_factor_auth():
        print("2FA failed. Access denied.")
        return

    if not device_check(allowed_devices):
        print("Device check failed. Access denied.")
        return

    print("Access granted. Proceeding with operations...")

    # Create Merkle root
    merkle_root = create_merkle_root(data_list)
    print(f"Merkle Root: {merkle_root}")

    # Update ledger
    update_ledger("Access granted to the file.")

    # Decrypt file before reading and writing
    if os.path.exists(key_file):
        key = load_key(key_file)
        decrypt_file(file_path, key)
        print("File decrypted successfully.")
    else:
        print("No key found, assuming file is not yet encrypted.")

    # Read existing data
    data_from_excel = read_excel_file(file_path)
    print("Data from Excel file:", data_from_excel)

    # Write new data
    new_data = [('New Data 1', 'New Data 2')]
    write_to_excel_file(file_path, new_data)
    print("New data written to Excel file.")

    # Encrypt file again
    if not os.path.exists(key_file):
        key = generate_key()
        save_key(key, key_file)
        print(f"Encryption key generated and saved to {key_file}")
    else:
        key = load_key(key_file)

    encrypt_file(file_path, key)
    print("File encrypted successfully.")

# Example usage
if __name__ == "__main__":
    allowed_devices = ['device1', 'device2', 'device3']
    data_list = ['data1', 'data2', 'data3']
    file_path = r'C:\Users\Ankita Ghosh\OneDrive\Desktop\PROJECT WORK\Bank_database_excel_sheet.xlsx'
    key_file = r'C:\Users\Ankita Ghosh\OneDrive\Desktop\PROJECT WORK\encryption_key.key'

    # Secure the file (encrypt, update Excel, manage key)
    secure_file(file_path, allowed_devices, data_list, key_file)

    # If you manually want to decrypt the file separately
    # Uncomment the following:
    # key = load_key(key_file)
    # decrypt_file(file_path, key)
    # print("File decrypted. You can now open it with Excel.")
